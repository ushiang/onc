import os
from django.conf import settings
from openpyxl import load_workbook


def migrate_mda():

    do_basic = True

    # MIGRATE AGENCIES
    from packages.budget.models import MDA

    # MIGRATE PERSONNEL RECORDS
    file_r = os.path.join(settings.BASE_DIR, "tmp/budget/data.xlsx")
    wb = load_workbook(file_r)
    ws = wb.active

    row = 1
    while do_basic:

        data = ws.cell("A%s" % row).value

        if data and data != "":
            code = data[:12]
            name = data[13:]

            print "Registering %s for %s" % (code, name)

            mda, c = MDA.pp.get_or_create(code=code, name=name)
            mda.code = code
            mda.name = name
            mda.save()

            print c

        else:
            break

        row += 1


def migrate_ncoa():

    do_basic = True

    # MIGRATE AGENCIES
    from packages.budget.models import Head, Chart

    # MIGRATE ACCOUNTS
    file_r = os.path.join(settings.BASE_DIR, "tmp/budget/data.xlsx")
    wb = load_workbook(file_r)
    ws = wb.get_sheet_by_name("NCOA")

    # PERSONNEL
    row = 2
    while do_basic:

        personnel = ws.cell("A%s" % row).value
        head = Head.pp.get(name="PERSONNEL")

        if personnel and personnel != "":
            code = personnel[:8]
            name = personnel[9:]

            print "Registering %s for %s - PERSONNEL" % (code, name)

            mda, c = Chart.pp.get_or_create(head=head, code=code, name=name)
            mda.code = code
            mda.name = name
            mda.save()

        else:
            break

        row += 1

    # OVERHEAD
    row = 2
    while do_basic:

        overhead = ws.cell("B%s" % row).value
        head = Head.pp.get(name="OVERHEAD")

        if overhead and overhead != "":
            code = overhead[:8]
            name = overhead[9:]

            print "Registering %s for %s - OVERHEAD" % (code, name)

            mda, c = Chart.pp.get_or_create(head=head, code=code, name=name)
            mda.code = code
            mda.name = name
            mda.save()

        else:
            break

        row += 1

    # CAPITAL
    row = 2
    while do_basic:

        capital = ws.cell("C%s" % row).value
        head = Head.pp.get(name="CAPITAL")

        if capital and capital != "":
            code = capital[:8]
            name = capital[9:]

            print "Registering %s for %s - CAPITAL" % (code, name)

            mda, c = Chart.pp.get_or_create(head=head, code=code, name=name)
            mda.code = code
            mda.name = name
            mda.save()

        else:
            break

        row += 1
