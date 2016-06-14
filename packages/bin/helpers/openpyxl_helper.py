from openpyxl.cell import column_index_from_string, get_column_letter, coordinate_from_string


def style_range(ws, cell_range, style=None):
    """
    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param style: An openpyxl Style object
    """

    start_cell, end_cell = cell_range.split(':')
    start_coord = coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = column_index_from_string(start_coord[0])
    end_coord = coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            col = get_column_letter(col_idx)
            ws.cell('%s%s' % (col, row)).style = style