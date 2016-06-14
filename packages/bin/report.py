import os
from django.conf import settings
from packages.bin.auth import Auth
from packages.bin.bin import format_alias
from packages.bin.lib import mkdir_if_not_exists
from packages.system.models import Module, Report as ReportModel


class DoReport(object):

    def __init__(self, request, module, data):
        self.request = request
        self.data = data
        self.module = Module.objects.get(alias=module)
        self.me = Auth.whoami(request, "users")

    def build(self):

        # process report and return to sheet for downloading
        from django.conf import settings
        from openpyxl import Workbook
        from openpyxl.styles import Style, PatternFill, Color, Font, Border, Side, borders, Alignment

        # styles
        border_style = Border(left=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                              right=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                              top=Side(border_style=borders.BORDER_THIN, color='FF000000'),
                              bottom=Side(border_style=borders.BORDER_THIN, color='FF000000'))

        pattern_fill = PatternFill(patternType='solid', fgColor=Color('FFCAE5EF'))

        alignment_style = Alignment(vertical='top', wrap_text=True, shrink_to_fit=False)

        font_header = Font(name='Tahoma', size=17)
        font_header_sub = Font(name='Tahoma', size=14)
        font_activity = Font(bold=True, name='Tahoma', size=12)
        font_basic = Font(name='Tahoma', size=12)

        s_header = Style(font=font_header, border=border_style)
        s_header_sub = Style(font=font_header_sub, fill=pattern_fill, border=border_style)
        s_activity = Style(font=font_activity, alignment=alignment_style, border=border_style)
        s_basic = Style(font=font_basic, alignment=alignment_style, border=border_style)

        # open workbook
        wb = Workbook()

        # create multiple sheets based on data
        sheet_index = 0
        for sheet in self.data["sheets"]:

            try:
                sheet_title = sheet["sheet_title"]
            except KeyError:
                sheet_title = sheet["title"]

            if sheet_index == 0:
                ws = wb.worksheets[0]
                ws.title = sheet_title
            else:
                ws = wb.create_sheet(sheet_title)

            sheet_index += 1

            # main title section
            col = "A"
            row = 1

            ws.cell('%s%s' % (col, row)).value = sheet["title"]
            ws.cell('%s%s' % (col, row)).style = s_header

            # column section
            col = "A"
            row += 1
            for h in sheet["columns"]:

                #ws.merge_cells(range_string='A1:B1')

                ws.cell('%s%s' % (col, row)).value = h
                ws.cell('%s%s' % (col, row)).style = s_header_sub

                col = chr(ord(col) + 1)

            # body section
            for row_data in sheet["body"]:

                row += 1
                col = "A"

                for x in row_data:

                    ws.cell('%s%s' % (col, row)).value = x
                    ws.cell('%s%s' % (col, row)).style = s_basic

                    col = chr(ord(col) + 1)

        mkdir_if_not_exists(settings.BASE_DIR+"/media/download/")

        file_location = settings.BASE_DIR + "/media/download/%s-%s.%s" % \
                                            (self.data["filename"], self.me.pk, self.data["ext"])

        wb.save(filename=file_location)

        # save into the report basket
        basket = Basket(self.request, self.module, self.data["filename"], file_location)
        basket.save()


class Basket(object):

    name = None
    file_url = None
    ext = None
    obj = None
    module = None
    request = None

    def __init__(self, request, module, name, file_url, ext=None):
        self.name = name
        self.file_url = file_url
        self.ext = ext or file_url.split('.')[len(file_url.split('.'))-1]
        self.module = module
        self.request = request

    def save(self):

        me = Auth.whoami(self.request, "users")

        # let us correct the file url so it doesn't come with the entire file path, who needs that
        self.file_url = self.file_url.replace(os.path.join(settings.MEDIA_ROOT, ''), '')
        if self.file_url[0:6] == 'media/':
            self.file_url.replace('media/', '', 1)

        # let us now delete any previous report before creating a new one, to preserve precious server space
        reports = ReportModel.objects.filter(name=self.name, mac=me)
        for report in reports:
            report.alt_delete()

        self.obj = ReportModel(module=self.module, name=self.name, file_url=self.file_url, ext=self.ext, mac=me)
        self.obj.save()